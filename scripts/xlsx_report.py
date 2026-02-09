#!/usr/bin/env python3
"""
PersonalShopper XLSX Report Generator
Generates formatted Excel reports from product search results.
Input: JSON via stdin or --input file
Output: XLSX file at --output path (default: ~/Desktop/shop_results_<timestamp>.xlsx)
"""

import json
import sys
import os
import subprocess
from datetime import datetime
from pathlib import Path

# Auto-bootstrap openpyxl with venv support for macOS Homebrew Python (PEP 668)
VENV_DIR = Path(__file__).parent / ".venv"

def _ensure_openpyxl():
    """Install openpyxl in a dedicated venv if not available system-wide."""
    try:
        import openpyxl
        return
    except ImportError:
        pass

    # Try --user install first
    try:
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "--user", "openpyxl", "-q"],
            stderr=subprocess.DEVNULL,
        )
        return
    except subprocess.CalledProcessError:
        pass

    # Fall back to venv
    if not VENV_DIR.exists():
        print("Creating venv for openpyxl...", file=sys.stderr)
        subprocess.check_call([sys.executable, "-m", "venv", str(VENV_DIR)])

    venv_pip = VENV_DIR / "bin" / "pip"
    subprocess.check_call([str(venv_pip), "install", "openpyxl", "-q"])

    # Add venv site-packages to sys.path
    import glob
    site_pkgs = glob.glob(str(VENV_DIR / "lib" / "python*" / "site-packages"))
    if site_pkgs:
        sys.path.insert(0, site_pkgs[0])

_ensure_openpyxl()

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# --- Styles ---
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2C2C31", end_color="2C2C31", fill_type="solid")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="2C2C31")
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="666666")
DATA_FONT = Font(name="Calibri", size=11)
LINK_FONT = Font(name="Calibri", size=11, color="1155CC", underline="single")
VERIFIED_FONT = Font(name="Calibri", size=11, color="0B6623")
UNVERIFIED_FONT = Font(name="Calibri", size=11, color="CC6600")
ROW_FILL_ALT = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
ROW_FILL_NORMAL = PatternFill(fill_type=None)
IN_STOCK_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
OUT_STOCK_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
UNVERIFIED_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)
HEADER_BORDER = Border(
    left=Side(style="thin", color="1A1A1A"),
    right=Side(style="thin", color="1A1A1A"),
    top=Side(style="thin", color="1A1A1A"),
    bottom=Side(style="medium", color="1A1A1A"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

# --- Column definitions ---
COLUMNS = [
    ("Nome", 35),
    ("Prezzo", 12),
    ("Prezzo Orig.", 12),
    ("Sconto %", 10),
    ("Sito", 20),
    ("URL", 15),
    ("Disponibilità", 15),
    ("Spedizione IT", 18),
    ("Note", 30),
    ("Verificato", 10),
]


def parse_args():
    import argparse
    parser = argparse.ArgumentParser(description="Generate XLSX shopping report")
    parser.add_argument("--input", "-i", help="Input JSON file (default: stdin)")
    parser.add_argument("--output", "-o", help="Output XLSX file path")
    parser.add_argument("--no-open", action="store_true", help="Don't auto-open the file")
    return parser.parse_args()


def load_data(input_path=None):
    if input_path:
        with open(input_path, "r", encoding="utf-8") as f:
            return json.load(f)
    else:
        return json.load(sys.stdin)


def auto_width(ws, col_idx, value, min_width=8, max_width=50):
    """Calculate column width based on content."""
    current = ws.column_dimensions[get_column_letter(col_idx)].width or 0
    if value is None:
        return
    length = min(len(str(value)) + 2, max_width)
    if length > current:
        ws.column_dimensions[get_column_letter(col_idx)].width = max(length, min_width)


def create_main_sheet(wb, data):
    ws = wb.active
    ws.title = "Risultati"

    # --- Title rows ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    title_cell = ws.cell(row=1, column=1, value=data.get("title", "Risultati Ricerca"))
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
    query_cell = ws.cell(row=2, column=1, value=f"Query: {data.get('query', 'N/D')}")
    query_cell.font = SUBTITLE_FONT

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(COLUMNS))
    date_cell = ws.cell(row=3, column=1, value=f"Generato il {data.get('date', datetime.now().strftime('%Y-%m-%d'))}")
    date_cell.font = SUBTITLE_FONT

    # --- Header row ---
    header_row = 5
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
        cell.border = HEADER_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # --- Data rows ---
    products = data.get("products", [])
    for row_idx, product in enumerate(products):
        row_num = header_row + 1 + row_idx
        is_alt = row_idx % 2 == 1
        fill = ROW_FILL_ALT if is_alt else ROW_FILL_NORMAL

        # Availability-based fill override
        disponibilita = str(product.get("disponibilita", "")).lower()
        if "out" in disponibilita or "esaurito" in disponibilita or "non disponibile" in disponibilita:
            fill = OUT_STOCK_FILL
        elif not product.get("verificato", True):
            fill = UNVERIFIED_FILL
        elif "in stock" in disponibilita or "disponibile" in disponibilita:
            if is_alt:
                fill = PatternFill(start_color="E0F2E0", end_color="E0F2E0", fill_type="solid")
            else:
                fill = IN_STOCK_FILL

        values = [
            product.get("nome", "N/D"),
            product.get("prezzo", "N/D"),
            product.get("prezzo_originale", ""),
            product.get("sconto", ""),
            product.get("sito", "N/D"),
            "Link",  # placeholder for hyperlink
            product.get("disponibilita", "N/D"),
            product.get("spedizione_it", "N/D"),
            product.get("note", ""),
            "✅" if product.get("verificato", False) else "⚠️",
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGNMENT

            # URL column: add hyperlink
            if col_idx == 6 and product.get("url"):
                cell.hyperlink = product["url"]
                cell.font = LINK_FONT
                cell.value = "Apri →"

            # Verified column styling
            if col_idx == 10:
                cell.alignment = CENTER_ALIGNMENT
                cell.font = VERIFIED_FONT if product.get("verificato") else UNVERIFIED_FONT

            auto_width(ws, col_idx, value)

    # --- Footer ---
    footer_row = header_row + len(products) + 2
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=len(COLUMNS))
    footer = ws.cell(
        row=footer_row,
        column=1,
        value=f"Generato da PersonalShopper Plugin — {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    )
    footer.font = Font(name="Calibri", size=9, italic=True, color="999999")

    # Freeze header
    ws.freeze_panes = f"A{header_row + 1}"

    return ws


def create_comparison_sheet(wb, data):
    """Create a comparison matrix sheet if type=comparison."""
    comparison = data.get("comparison_extra")
    if not comparison:
        return

    ws = wb.create_sheet("Confronto")

    ws.cell(row=1, column=1, value="Confronto Prodotti").font = TITLE_FONT

    matrix = comparison.get("matrix", [])
    if not matrix:
        return

    # Headers
    headers = ["Criterio", "Peso"] + [p.get("nome", f"Prodotto {i+1}") for i, p in enumerate(data.get("products", []))]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = CENTER_ALIGNMENT

    # Data
    for row_idx, row_data in enumerate(matrix):
        row_num = 4 + row_idx
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGNMENT if col_idx > 1 else WRAP_ALIGNMENT

    # Winner
    winner = comparison.get("winner")
    if winner:
        winner_row = 4 + len(matrix) + 1
        ws.cell(row=winner_row, column=1, value="🏆 VINCITORE").font = Font(
            name="Calibri", size=12, bold=True, color="0B6623"
        )
        ws.cell(row=winner_row, column=2, value=winner).font = Font(
            name="Calibri", size=12, bold=True
        )


def generate_csv_fallback(data, output_path):
    """Fallback to CSV if openpyxl fails."""
    import csv

    csv_path = output_path.with_suffix(".csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([col[0] for col in COLUMNS])
        for product in data.get("products", []):
            writer.writerow([
                product.get("nome", ""),
                product.get("prezzo", ""),
                product.get("prezzo_originale", ""),
                product.get("sconto", ""),
                product.get("sito", ""),
                product.get("url", ""),
                product.get("disponibilita", ""),
                product.get("spedizione_it", ""),
                product.get("note", ""),
                "SI" if product.get("verificato") else "NO",
            ])
    return csv_path


def main():
    args = parse_args()

    # Load data
    try:
        data = load_data(args.input)
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON input: {e}", file=sys.stderr)
        sys.exit(1)
    except FileNotFoundError:
        print(f"Error: File not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    # Determine output path
    if args.output:
        output_path = Path(args.output)
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path.home() / "Desktop" / f"shop_results_{timestamp}.xlsx"

    # Generate XLSX
    try:
        wb = openpyxl.Workbook()
        create_main_sheet(wb, data)

        if data.get("type") == "comparison":
            create_comparison_sheet(wb, data)

        wb.save(str(output_path))
        print(f"Report saved: {output_path}")

        # Auto-open on macOS
        if not args.no_open and sys.platform == "darwin":
            subprocess.run(["open", str(output_path)], check=False)

    except Exception as e:
        print(f"XLSX generation failed ({e}), falling back to CSV...", file=sys.stderr)
        csv_path = generate_csv_fallback(data, output_path)
        print(f"CSV fallback saved: {csv_path}")
        if not args.no_open and sys.platform == "darwin":
            subprocess.run(["open", str(csv_path)], check=False)


if __name__ == "__main__":
    main()
